"""momiji-mcp — MomijiStore OS MCP Server

公式SDK(mcp)の標準サーバーAPIのみを使用する。独自実装はしない。

Phase3 STEP4: health_check
Phase3 STEP5-1: search_products(商品マスター検索・読み取り専用)
Phase3 STEP6:   Intelligence Layer(判断・施策・結果の記録)

⚠️ Excelに対しては読み取り専用。書き込み・更新・削除は一切行わない。
   NAS上のExcelは「MCP検索用の読み取り専用スナップショット」であり正本ではない。
   正本は Mac側 SourceData/商品マスター_単品_v1.0.xlsx。

   書き込みが許されるのは intelligence スキーマの2表のみ。
   これらはExcelに対応物を持たない「DBで生まれるデータ」であり、
   正本が二重化しない(BL-5)。詳細は src/intelligence.py 冒頭。
"""

from __future__ import annotations

import logging
import os
import secrets
import threading
import time
from pathlib import Path
from typing import Any, Protocol

import openpyxl
from mcp.server import MCPServer

import intelligence
from intelligence import IntelligenceError

# starlette / uvicorn は mcp SDK の依存として既に入っている。
# 認証ミドルウェアと /health を足すために直接importする(新規の追加依存ではない)。
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from starlette.responses import JSONResponse

SERVICE_NAME = "momiji-mcp"

# --- 認証 -----------------------------------------------------------------
#  OAuth 2.1 と同じ Bearer トークン方式を使う(MCP仕様が定める形)。
#  /mcp 配下は全て認証必須。MCPのツールはすべて /mcp を通るため、
#  今後ツールを追加しても自動的に認証が適用される(個別対応は不要)。
#  /health のみ認証不要(監視用)。
#
#  MOMIJI_API_KEYS はカンマ区切りで複数指定できる(ローテーション用)。
#  新旧の鍵を並べておき、クライアント移行後に旧鍵を消す運用を想定する。
#  MOMIJI_API_KEY(単数)は旧設定との互換のためのフォールバック。
def _load_api_keys() -> list[str]:
    raw = os.environ.get("MOMIJI_API_KEYS", "") or os.environ.get("MOMIJI_API_KEY", "")
    return [k for k in (part.strip() for part in raw.split(",")) if k]


API_KEYS = _load_api_keys()
PUBLIC_PATHS = frozenset({"/health"})
# 認証失敗時に返すヘッダー。MCP仕様は 401 + WWW-Authenticate: Bearer を求める。
WWW_AUTHENTICATE = 'Bearer error="invalid_token"'

# レート制限(IP単位)。メモリ上の固定ウィンドウ方式。
RATE_LIMIT_PER_MINUTE = int(os.getenv("MOMIJI_RATE_LIMIT_PER_MINUTE", "100"))
RATE_LIMIT_WINDOW_SEC = 60

# SDKのDNSリバインディング対策。既定では localhost しか許可されず、
# LAN内のIP(192.168.0.8:8000)で叩くと 421 Misdirected Request になる。
# 到達を許すホストを明示する。既定はNASのLAN側アドレスとループバック。
ALLOWED_HOSTS = [
    h.strip()
    for h in os.getenv(
        "MOMIJI_ALLOWED_HOSTS", "192.168.0.8:8000,localhost:*,127.0.0.1:*"
    ).split(",")
    if h.strip()
]


def key_id(token: str) -> str:
    """ログ用の識別子。トークン全文は絶対に出さず、先頭8文字だけを使う。"""
    return token[:8]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(SERVICE_NAME)

mcp = MCPServer(SERVICE_NAME)

# 読み取り専用マウント (:ro) されたスナップショット
PRODUCTS_XLSX = os.environ.get(
    "MOMIJI_PRODUCTS_XLSX", "/app/data/products/商品マスター_単品_v1.0.xlsx"
)
# 起動時点で「実際にどのパスを見るか」をログへ残す。
# 環境変数で上書きされていた場合、ここで判明する。
logger.info(
    "起動 参照パス=%s (環境変数指定=%s)",
    PRODUCTS_XLSX,
    "あり" if os.environ.get("MOMIJI_PRODUCTS_XLSX") else "なし(既定値)",
)

SHEET_MASTER = "商品マスター"
SHEET_LISTING = "出品テーブル"

# 商品マスター シートの列位置と、想定する列見出し(変更検知に使う)
PM_INTERNAL_ID, PM_JAN, PM_NAME = 0, 1, 2
PM_COST, PM_CHANNEL, PM_STATUS = 4, 5, 6
PM_HEADERS = {
    PM_INTERNAL_ID: "内部管理ID",
    PM_JAN: "JAN",
    PM_NAME: "商品名",
    PM_COST: "標準原価",
    PM_CHANNEL: "販売チャネル",
    PM_STATUS: "販売ステータス",
}

# 出品テーブル シートの列位置と、想定する列見出し
LT_INTERNAL_ID, LT_RAKUTEN_SKU, LT_ASIN = 1, 4, 5
LT_HEADERS = {
    LT_INTERNAL_ID: "内部管理ID",
    LT_RAKUTEN_SKU: "楽天SKU",
    LT_ASIN: "ASIN",
}

# 1回の検索で返す最大件数(応答の肥大を防ぐ)
# .env / docker-compose 側で変更できる。既定は 50 / 500。
DEFAULT_LIMIT = int(os.getenv("MOMIJI_DEFAULT_LIMIT", "50"))
MAX_LIMIT = int(os.getenv("MOMIJI_MAX_LIMIT", "500"))

# メモリキャッシュ(初回アクセス時にロード。ファイル更新で自動再ロード)
_cache: list[dict[str, Any]] = []
_cache_mtime: float | None = None
_cache_lock = threading.Lock()


class ProductDataError(Exception):
    """商品マスターを読めなかった。メッセージは利用者へそのまま返せる内容にする。"""


def _path_diagnostics(exc: BaseException) -> str:
    """アクセス失敗時に、原因を特定できるだけの情報を組み立てる。

    例外を握り潰すと調査ができなくなるため、実際の例外型・メッセージと、
    参照パス・存在確認・親ディレクトリの内容を必ず添える。
    ここで扱うのはコンテナ内部のパスであり、秘密情報は含まない。
    """
    target = Path(PRODUCTS_XLSX)
    parent = target.parent
    try:
        entries: Any = sorted(os.listdir(parent))
    except OSError as list_exc:
        entries = f"<一覧取得不可: {type(list_exc).__name__}: {list_exc}>"
    return (
        f"商品マスターへアクセスできません({type(exc).__name__}: {exc})。"
        f" 参照パス={PRODUCTS_XLSX}"
        f" / exists={target.exists()} is_file={target.is_file()}"
        f" / 親={parent} exists={parent.exists()} is_dir={parent.is_dir()}"
        f" / 親の内容={entries}"
    )


def _verify_headers(sheet, expected: dict[int, str], sheet_name: str) -> None:
    """列見出しが想定どおりかを確認する。

    列構成が変わったまま読み進めると誤った値を返してしまう。
    黙って間違えるより、明確に失敗させる(BL-7 記録なき変更を認めない)。
    """
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header is None:
        raise ProductDataError(f"シート「{sheet_name}」が空です。")
    for index, name in expected.items():
        actual = header[index] if index < len(header) else None
        if actual != name:
            raise ProductDataError(
                f"シート「{sheet_name}」の列構成が想定と異なります"
                f"(列{index + 1}: 期待「{name}」/ 実際「{actual}」)。"
                "商品マスターの形式変更が疑われます。"
            )


def _load_products() -> list[dict[str, Any]]:
    """Excelを読み取り専用で開き、商品マスターと出品テーブルを突き合わせて返す。

    ASIN・楽天SKUは出品テーブル側にあり、1商品が複数の出品行を持つ場合がある
    (楽天とAmazonの両方など)。値が存在しない場合は推測せず None を返す。
    """
    logger.info("商品マスターを開きます path=%s", PRODUCTS_XLSX)
    try:
        wb = openpyxl.load_workbook(PRODUCTS_XLSX, read_only=True, data_only=True)
    except (FileNotFoundError, PermissionError, OSError) as exc:
        logger.exception("商品マスターを開けません path=%s", PRODUCTS_XLSX)
        raise ProductDataError(_path_diagnostics(exc)) from exc
    except Exception as exc:
        # 破損ファイル・非対応形式など
        logger.exception("商品マスターの読み込みに失敗 path=%s", PRODUCTS_XLSX)
        raise ProductDataError(
            f"商品マスターのファイルを開けませんでした({type(exc).__name__}: {exc})。"
            "ファイルが破損していないか確認してください。"
        ) from exc

    try:
        for sheet_name in (SHEET_MASTER, SHEET_LISTING):
            if sheet_name not in wb.sheetnames:
                raise ProductDataError(
                    f"シート「{sheet_name}」が見つかりません。"
                    f"存在するシート: {', '.join(wb.sheetnames)}"
                )
        _verify_headers(wb[SHEET_LISTING], LT_HEADERS, SHEET_LISTING)
        _verify_headers(wb[SHEET_MASTER], PM_HEADERS, SHEET_MASTER)

        listings: dict[str, dict[str, Any]] = {}
        for row in wb[SHEET_LISTING].iter_rows(min_row=2, values_only=True):
            pid = row[LT_INTERNAL_ID]
            if not pid:
                continue
            entry = listings.setdefault(pid, {"asin": None, "rakuten_sku": None})
            # 最初に見つかった非空の値を採用する
            if entry["asin"] is None and row[LT_ASIN]:
                entry["asin"] = str(row[LT_ASIN])
            if entry["rakuten_sku"] is None and row[LT_RAKUTEN_SKU]:
                entry["rakuten_sku"] = str(row[LT_RAKUTEN_SKU])

        products = []
        seen: set[str] = set()
        duplicates = 0
        for row in wb[SHEET_MASTER].iter_rows(min_row=2, values_only=True):
            pid = row[PM_INTERNAL_ID]
            if not pid:
                continue
            if pid in seen:
                duplicates += 1
            seen.add(pid)
            listing = listings.get(pid, {})
            products.append(
                {
                    "内部管理ID": pid,
                    "商品名": row[PM_NAME],
                    "JAN": str(row[PM_JAN]) if row[PM_JAN] is not None else None,
                    "ASIN": listing.get("asin"),
                    "楽天SKU": listing.get("rakuten_sku"),
                    "標準原価": row[PM_COST],
                    "販売チャネル": row[PM_CHANNEL],
                    "販売ステータス": row[PM_STATUS],
                }
            )
        if duplicates:
            # 正本側の問題なので落とさず警告する。該当行は全件返す。
            logger.warning("内部管理IDの重複を検出しました 件数=%d", duplicates)
        return products
    finally:
        wb.close()


def _ensure_loaded() -> list[dict[str, Any]]:
    """キャッシュを返す。未ロード、またはファイルが更新されていれば再ロードする。

    同時アクセスでの二重ロードを避けるためロックを取る。
    """
    global _cache, _cache_mtime
    try:
        mtime = os.path.getmtime(PRODUCTS_XLSX)
    except OSError as exc:
        # 例外を握り潰さない。実際の型・メッセージとパス診断を必ず残す。
        logger.exception("商品マスターへアクセスできません path=%s", PRODUCTS_XLSX)
        raise ProductDataError(_path_diagnostics(exc)) from exc

    with _cache_lock:
        if _cache_mtime != mtime:
            _cache = _load_products()
            _cache_mtime = mtime
            logger.info("商品マスターをロードしました 件数=%d", len(_cache))
        return _cache


@mcp.tool()
def health_check() -> dict:
    """MCPサーバーが応答することを確認する。"""
    return {"status": "ok", "service": SERVICE_NAME}


@mcp.tool()
def search_products(
    name: str | None = None,
    jan: str | None = None,
    internal_id: str | None = None,
    asin: str | None = None,
    rakuten_sku: str | None = None,
    limit: int = DEFAULT_LIMIT,
) -> dict:
    """商品マスターを検索する(読み取り専用)。

    商品名は部分一致、それ以外は完全一致。複数指定した場合はAND条件。
    条件を1つも指定しない場合は全件返さず、エラーを返す。

    Args:
        name: 商品名(部分一致)
        jan: JANコード(完全一致)
        internal_id: 内部管理ID(完全一致)
        asin: ASIN(完全一致)
        rakuten_sku: 楽天SKU(完全一致)
        limit: 返却する最大件数(既定50・上限500)。超過分は truncated=true で示す
    """
    started = time.perf_counter()
    conditions = {
        "name": name,
        "jan": jan,
        "internal_id": internal_id,
        "asin": asin,
        "rakuten_sku": rakuten_sku,
    }
    given = {k: v for k, v in conditions.items() if v}

    if not given:
        logger.info("search_products 条件なし — 検索を拒否しました")
        return {
            "error": "検索条件を1つ以上指定してください "
            "(name / jan / internal_id / asin / rakuten_sku)。"
            "全件返却は行いません。",
            "results": [],
            "count": 0,
        }

    try:
        products = _ensure_loaded()
    except ProductDataError as exc:
        logger.error("search_products 読み込み失敗: %s", exc)
        return {"error": str(exc), "results": [], "count": 0, "matched": 0}

    effective_limit = max(1, min(int(limit), MAX_LIMIT))
    results = []
    matched = 0
    for p in products:
        if name and (not p["商品名"] or name not in str(p["商品名"])):
            continue
        if jan and p["JAN"] != str(jan):
            continue
        if internal_id and p["内部管理ID"] != internal_id:
            continue
        if asin and p["ASIN"] != str(asin):
            continue
        if rakuten_sku and p["楽天SKU"] != str(rakuten_sku):
            continue
        matched += 1
        if len(results) < effective_limit:
            results.append(p)

    elapsed_ms = (time.perf_counter() - started) * 1000
    logger.info(
        "search_products 条件=%s 該当=%d件 返却=%d件 処理時間=%.1fms",
        given,
        matched,
        len(results),
        elapsed_ms,
    )
    return {
        "results": results,
        "count": len(results),
        "matched": matched,
        "truncated": matched > len(results),
        "elapsed_ms": round(elapsed_ms, 1),
    }


# --- Intelligence Layer(知識層)-------------------------------------------
#  判断・施策・結果を記録する。分析だけではAIは成長しないため、
#  「何を判断し / なぜそう判断し / 結果どうなったか」を対で残す。
#
#  ⚠️ 記録は追記専用。書き換え・削除はできない(DBトリガーでも拒否される)。
#     訂正は supersedes に旧IDを入れた新しい記録で行う。
#
#  各ツールのdocstringは、AIがこの層を正しく使うための唯一の説明になる。
#  実装を変えるときはdocstringも必ず合わせる。


def _intel(operation: str, func, /, **kwargs) -> dict:
    """Intelligence Layer の共通の呼び出し口。

    例外は握り潰さず、必ずログへ残したうえで利用者に読めるメッセージを返す。
    """

    def failure(message: str) -> dict:
        # 記録系は「記録できていない」ことが伝わらないと危険なので明示する。
        # 参照系は search_products と同じ形(results/count)に揃える。
        if operation.startswith("record_"):
            return {"error": message, "recorded": False}
        return {"error": message, "results": [], "count": 0, "matched": 0}

    try:
        return func(**kwargs)
    except IntelligenceError as exc:
        logger.error("%s 失敗: %s", operation, exc)
        return failure(str(exc))
    except Exception as exc:
        logger.exception("%s で予期しないエラー", operation)
        return failure(f"{operation} に失敗しました({type(exc).__name__}: {exc})。")


@mcp.tool()
def record_decision(
    decision_type: str,
    action: str,
    action_kind: str,
    reason: str,
    decided_by: str,
    subject_type: str | None = None,
    subject_id: str | None = None,
    subject_label: str | None = None,
    alternatives: str | None = None,
    expected: str | None = None,
    expected_metric: str | None = None,
    expected_value: float | None = None,
    review_due: str | None = None,
    decided_at: str | None = None,
    proposed_by: str | None = None,
    business_logic: str | None = None,
    supersedes: int | None = None,
    note: str | None = None,
) -> dict:
    """経営判断を1件記録する(追記専用)。

    仕入・価格改定・広告・在庫など、何かを決めたら必ずここへ残す。
    **「今回は変更しない」「この候補は見送る」も判断であり、記録する**
    (BL-11)。却下理由が残らないと、AIは同じ提案を繰り返す。

    Args:
        decision_type: 判断の型。Decision_Catalog.md のID(例: DEC-SAL-01)
        action: 実際に取った施策(例:「1,980円へ値下げ」「価格を据え置いた」)
        action_kind: changed(変更した)/ unchanged(変更しなかった)/ rejected(却下した)
        reason: なぜそう判断したか。**この層の中心。曖昧に書かない**
        decided_by: 判断した人の名前。AIは指定できない(AI Constitution 第1条)
        subject_type: 対象の種類(product / campaign / supplier / system 等)
        subject_id: 対象の識別子(内部管理ID / ASIN / 楽天SKU / キャンペーンID)
        subject_label: 対象の名称(人が読むため)
        alternatives: 検討したが採らなかった案
        expected: 期待した結果(定性)
        expected_metric: 検証する指標名(利益額 / TACOS / ROAS / 回転率 等)
        expected_value: 期待値(数値)
        review_due: いつ結果を確認するか(YYYY-MM-DD)。**未設定にしない**
        decided_at: 判断日時。省略時は現在時刻
        proposed_by: 提案者。AI提案なら 'ai:claude' 等
        business_logic: 根拠にしたBL番号(例: "BL-3,BL-4")
        supersedes: 過去の記録を訂正する場合、その decision_id
        note: 補足
    """
    return _intel(
        "record_decision",
        intelligence.record_decision,
        decision_type=decision_type,
        action=action,
        action_kind=action_kind,
        reason=reason,
        decided_by=decided_by,
        subject_type=subject_type,
        subject_id=subject_id,
        subject_label=subject_label,
        alternatives=alternatives,
        expected=expected,
        expected_metric=expected_metric,
        expected_value=expected_value,
        review_due=review_due,
        decided_at=decided_at,
        proposed_by=proposed_by,
        business_logic=business_logic,
        supersedes=supersedes,
        note=note,
    )


@mcp.tool()
def record_outcome(
    decision_id: int,
    assessment: str,
    summary: str,
    measured_by: str,
    metric: str | None = None,
    actual_value: float | None = None,
    period_start: str | None = None,
    period_end: str | None = None,
    learning: str | None = None,
    measured_at: str | None = None,
    note: str | None = None,
) -> dict:
    """判断の結果を記録する(追記専用)。同じ判断に何度でも記録してよい。

    **記録した判断は必ず結果まで残す。** やりっぱなしにすると、
    AIは自分の提案が正しかったかを学習できない。

    Args:
        decision_id: 対象の判断ID(record_decision の戻り値)
        assessment: success / partial / failure / unclear。
            他要因が大きく切り分けられない場合は無理に成否を付けず unclear
        summary: 何が起きたか
        measured_by: 測定・評価した人の名前
        metric: 測った指標名(判断時の expected_metric と揃える)
        actual_value: 実績値
        period_start: 測定対象期間の開始(YYYY-MM-DD)
        period_end: 測定対象期間の終了(YYYY-MM-DD)
        learning: 次に活かす学び。**なぜそうなったかまで書く**
        measured_at: 測定日時。省略時は現在時刻
        note: 補足
    """
    return _intel(
        "record_outcome",
        intelligence.record_outcome,
        decision_id=decision_id,
        assessment=assessment,
        summary=summary,
        measured_by=measured_by,
        metric=metric,
        actual_value=actual_value,
        period_start=period_start,
        period_end=period_end,
        learning=learning,
        measured_at=measured_at,
        note=note,
    )


@mcp.tool()
def search_decisions(
    decision_type: str | None = None,
    subject_id: str | None = None,
    action_kind: str | None = None,
    decided_by: str | None = None,
    since: str | None = None,
    until: str | None = None,
    keyword: str | None = None,
    with_outcomes: bool = True,
    limit: int = intelligence.DEFAULT_LIMIT,
) -> dict:
    """過去の判断と、その結果を検索する。

    **新しい施策を提案する前に、まずここを引く。** 同じ対象を過去に
    どう判断し、それがどうなったかを見ずに提案してはならない。
    action_kind="rejected" を引けば「過去に見送った理由」が分かる。

    Args:
        decision_type: 判断の型で絞る(例: DEC-SAL-01)
        subject_id: 対象の識別子で絞る(内部管理ID / ASIN 等)
        action_kind: changed / unchanged / rejected で絞る
        decided_by: 判断者で絞る
        since: この日時以降(YYYY-MM-DD または ISO8601)
        until: この日時以前
        keyword: 施策・理由・対象名の部分一致
        with_outcomes: 結果も併せて返すか(既定 true)
        limit: 返却する最大件数(既定50・上限500)
    """
    return _intel(
        "search_decisions",
        intelligence.search_decisions,
        decision_type=decision_type,
        subject_id=subject_id,
        action_kind=action_kind,
        decided_by=decided_by,
        since=since,
        until=until,
        keyword=keyword,
        with_outcomes=with_outcomes,
        limit=limit,
    )


@mcp.tool()
def list_pending_reviews(
    as_of: str | None = None,
    only_due: bool = True,
    limit: int = intelligence.DEFAULT_LIMIT,
) -> dict:
    """結果がまだ記録されていない判断の一覧を返す。

    「やりっぱなし」を可視化する。ここが溜まっている状態は、
    施策を打ちっぱなしで検証していないことを意味する。

    Args:
        as_of: 基準日(YYYY-MM-DD)。省略時は今日
        only_due: true なら確認期限が到来したもの(と期限未設定のもの)だけ返す
        limit: 返却する最大件数(既定50・上限500)
    """
    return _intel(
        "list_pending_reviews",
        intelligence.list_pending_reviews,
        as_of=as_of,
        only_due=only_due,
        limit=limit,
    )


def _extract_bearer_token(request: Request) -> str:
    """Authorization ヘッダーから Bearer トークンを取り出す。

    形式が違えば空文字を返す。scheme の判定は大文字小文字を区別しない
    (RFC 6750 に従う)。
    """
    header = request.headers.get("Authorization", "")
    scheme, _, token = header.partition(" ")
    if scheme.lower() != "bearer":
        return ""
    return token.strip()


class TokenVerifier(Protocol):
    """トークン検証の共通インターフェース。

    OAuth2 / JWT へ移行する際は、このインターフェースを満たす
    OAuthTokenVerifier / JWTTokenVerifier を実装して差し替える。
    ミドルウェア本体・レスポンス形式・クライアント設定は変更不要。
    """

    def verify(self, token: str) -> str | None:
        """有効ならログ用の識別子(key_id)を、無効なら None を返す。"""
        ...


class StaticTokenVerifier:
    """環境変数で与えた静的トークンと突き合わせる。

    複数トークンを許可し、どれか1つに一致すれば有効とする(ローテーション用)。
    比較は secrets.compare_digest による定数時間比較。
    """

    def __init__(self, keys: list[str]) -> None:
        self._keys = list(keys)

    def verify(self, token: str) -> str | None:
        if not token:
            return None
        for key in self._keys:
            if secrets.compare_digest(token, key):
                return key_id(key)
        return None


class RateLimiter:
    """IP単位の固定ウィンドウ方式レート制限(メモリ保持)。"""

    def __init__(self, limit: int, window_sec: int) -> None:
        self._limit = limit
        self._window = window_sec
        self._hits: dict[str, tuple[float, int]] = {}
        self._lock = threading.Lock()

    def allow(self, client_ip: str) -> bool:
        now = time.monotonic()
        with self._lock:
            window_start, count = self._hits.get(client_ip, (now, 0))
            if now - window_start >= self._window:
                window_start, count = now, 0
            count += 1
            self._hits[client_ip] = (window_start, count)
            # 古いエントリを掃除してメモリの増え続けを防ぐ
            if len(self._hits) > 1024:
                self._hits = {
                    ip: v
                    for ip, v in self._hits.items()
                    if now - v[0] < self._window
                }
            return count <= self._limit


class BearerAuthMiddleware(BaseHTTPMiddleware):
    """レート制限とBearerトークン認証を行う。

    MCPのツール呼び出しはすべて /mcp を通るため、この1箇所で
    既存・将来を問わず全ツールが保護される。/health のみ認証不要。
    """

    def __init__(self, app, verifier: TokenVerifier, limiter: RateLimiter) -> None:
        super().__init__(app)
        self._verifier = verifier
        self._limiter = limiter

    async def dispatch(self, request: Request, call_next):
        # レート制限は認証前に適用する(総当たり攻撃も抑止するため)
        client_ip = request.client.host if request.client else "unknown"
        if not self._limiter.allow(client_ip):
            logger.warning(
                "レート制限 method=%s path=%s ip=%s",
                request.method,
                request.url.path,
                client_ip,
            )
            return JSONResponse(
                {
                    "error": "rate_limit_exceeded",
                    "error_description": (
                        f"1分あたり {RATE_LIMIT_PER_MINUTE} 回までです。"
                    ),
                },
                status_code=429,
                headers={"Retry-After": str(RATE_LIMIT_WINDOW_SEC)},
            )

        if request.url.path in PUBLIC_PATHS:
            return await call_next(request)

        matched = self._verifier.verify(_extract_bearer_token(request))
        if matched is None:
            # トークンの値は一切記録しない。method と path のみ。
            logger.warning(
                "認証失敗 method=%s path=%s", request.method, request.url.path
            )
            return JSONResponse(
                {
                    "error": "invalid_token",
                    "error_description": "Authentication required",
                },
                status_code=401,
                headers={"WWW-Authenticate": WWW_AUTHENTICATE},
            )

        # どの鍵で認証されたかを追えるようにする(先頭8文字のみ)
        logger.info(
            "認証成功 method=%s path=%s key_id=%s",
            request.method,
            request.url.path,
            matched,
        )
        return await call_next(request)


async def health_endpoint(request: Request) -> JSONResponse:
    """認証不要の監視用エンドポイント。データには一切触れない。"""
    return JSONResponse({"status": "ok", "service": SERVICE_NAME})


if __name__ == "__main__":
    # トークン未設定なら起動しない(認証なしで公開される事故を防ぐ)
    if not API_KEYS:
        raise SystemExit("MOMIJI_API_KEY が設定されていません")

    import uvicorn
    from mcp.server.transport_security import TransportSecuritySettings

    app = mcp.streamable_http_app(
        transport_security=TransportSecuritySettings(
            allowed_hosts=ALLOWED_HOSTS,
            # Originはブラウザからの利用を想定していないため許可しない
            allowed_origins=[],
        )
    )
    app.add_middleware(
        BearerAuthMiddleware,
        verifier=StaticTokenVerifier(API_KEYS),
        limiter=RateLimiter(RATE_LIMIT_PER_MINUTE, RATE_LIMIT_WINDOW_SEC),
    )
    app.add_route("/health", health_endpoint, methods=["GET"])

    logger.info(
        "起動します 認証=Bearer 有効鍵=%d本(key_id=%s) レート制限=%d回/分"
        " 認証不要=%s 許可ホスト=%s",
        len(API_KEYS),
        ",".join(key_id(k) for k in API_KEYS),
        RATE_LIMIT_PER_MINUTE,
        sorted(PUBLIC_PATHS),
        ALLOWED_HOSTS,
    )
    # ポートはコンテナ内部で待ち受け、公開範囲は compose の ports で制御する。
    uvicorn.run(app, host="0.0.0.0", port=8000)
