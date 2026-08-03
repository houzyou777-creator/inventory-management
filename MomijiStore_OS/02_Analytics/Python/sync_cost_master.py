# -*- coding: utf-8 -*-
"""sync_cost_master.py — 原価一元管理の同期ツール

商品マスターを「原価の正」とするための同期スクリプト。
KPI管理シートの月次シートと商品マスターを突き合わせる。

使い方:
    python3 sync_cost_master.py check 7月            # 突合レポートのみ(変更なし)
    python3 sync_cost_master.py register 7月         # 未登録商品をマスターへ追加登録
    python3 sync_cost_master.py adopt 7月            # 原価不一致をKPI側の値でマスター更新
                                                     # (ユーザーが明示承認した場合のみ実行すること)
    python3 sync_cost_master.py push                 # 商品マスター→在庫集計ツール原価マスターへ配信

    python3 sync_cost_master.py check-amazon 7月     # Amazon KPIシートとの突合レポート
    python3 sync_cost_master.py register-amazon 7月  # Amazon側の未登録商品をマスターへ追加登録
    python3 sync_cost_master.py adopt-amazon 7月     # Amazon側の原価不一致をKPI側でマスター更新(要承認)

注意: push は openpyxl ではなく Excel(AppleScript)経由で書き込む。
xlsm を openpyxl で保存するとボタン描画(drawing1.xml)が失われるため。
GS移行時はこの層を Apps Script の直接参照に置き換える。

設計方針(GS移行を見据えて):
- 照合キーは「楽天商品管理番号 × 楽天SKU」のペア(SKU単独照合はNG。使い回しがあるため)
- マスターの既存原価は絶対に上書きしない。不一致は一覧表示してユーザー判断に委ねる
- 追加登録行は備考に出所を残し、後から見分けられるようにする
"""
import re
import sys
from datetime import date

from openpyxl import load_workbook

BASE = '/Users/hide0726/Desktop/Claude Code/MomijiStore_OS'
MASTER_FILE = BASE + '/01_InventoryManagement/SourceData/商品マスター_単品_v1.0.xlsx'
KPI_FILE = BASE + '/02_Analytics/SourceData/楽天運営 KPI管理シート.xlsx'

SH_MASTER = '商品マスター'
SH_LISTING = '出品テーブル'

# 商品マスター列 (1始まり)
PM_ID, PM_JAN, PM_NAME, PM_TYPE, PM_COST, PM_CH, PM_STATUS, PM_REG, PM_UPD, PM_NOTE = range(1, 11)
# 出品テーブル列
LT_ID, LT_PID, LT_CH, LT_RCTRL, LT_RSKU, LT_ASIN, LT_ASKU, LT_PRICE, LT_STATUS, LT_REG, LT_NOTE = range(1, 12)


def norm(v):
    return str(v).strip().upper() if v is not None and str(v).strip() else ''


def load_master():
    """マスターを読み、(管理番号,SKU)ペア→内部管理ID と ID→標準原価 を返す"""
    wb = load_workbook(MASTER_FILE, data_only=True)
    cost_by_pid = {}
    for r in wb[SH_MASTER].iter_rows(min_row=2, values_only=True):
        if r[0]:
            cost_by_pid[str(r[0])] = r[PM_COST - 1]
    pair2pid, ctrl2pid = {}, {}
    for r in wb[SH_LISTING].iter_rows(min_row=2, values_only=True):
        pid, rctrl, rsku = r[LT_PID - 1], r[LT_RCTRL - 1], r[LT_RSKU - 1]
        if not pid or not rctrl:
            continue
        pair2pid.setdefault((norm(rctrl), norm(rsku)), str(pid))
        ctrl2pid.setdefault(norm(rctrl), str(pid))
    return cost_by_pid, pair2pid, ctrl2pid


def load_kpi_month(month_sheet):
    """KPI月次シートから (管理番号, SKU, 商品名, 商品番号, 仕入値) を行順で返す"""
    wb = load_workbook(KPI_FILE, data_only=True)
    ws = wb[month_sheet]
    out = []
    for row in range(8, ws.max_row + 1):
        ctrl = ws.cell(row, 3).value
        if ctrl is None:          # データ末尾(合計行の手前)まで
            break
        out.append({
            'ctrl': str(ctrl).strip(),
            'sku': str(ws.cell(row, 5).value).strip() if ws.cell(row, 5).value else '',
            'name': str(ws.cell(row, 1).value or '').strip(),
            'pn': str(ws.cell(row, 4).value or '').strip(),
            'cost': ws.cell(row, 12).value,
        })
    return out


def classify(rows, cost_by_pid, pair2pid, ctrl2pid):
    """KPI行をマスターと突合して 一致/不一致/未登録 に分類(ペア単位で重複排除)"""
    match, conflict, missing, seen = [], [], [], set()
    for r in rows:
        key = (norm(r['ctrl']), norm(r['sku']))
        if key in seen:
            continue
        seen.add(key)
        pid = pair2pid.get(key) or ctrl2pid.get(key[0])
        if pid is None:
            missing.append(r)
        else:
            mc = cost_by_pid.get(pid)
            if mc is not None and r['cost'] is not None and float(mc) == float(r['cost']):
                match.append(r)
            else:
                conflict.append((r, pid, mc))
    return match, conflict, missing


def extract_jan(pn):
    """商品番号欄から13桁JANらしき先頭数字列を取り出す(なければ空)"""
    m = re.match(r'(\d{13})', pn)
    return m.group(1) if m else ''


def register_missing(month_sheet, missing):
    """未登録商品を商品マスター+出品テーブルへ追加登録する(既存行は変更しない)"""
    wb = load_workbook(MASTER_FILE)
    wm, wl = wb[SH_MASTER], wb[SH_LISTING]
    pids = [str(r[0]) for r in wm.iter_rows(min_row=2, values_only=True) if r[0]]
    cids = [str(r[0]) for r in wl.iter_rows(min_row=2, values_only=True) if r[0]]
    pn_num = max(int(p[1:]) for p in pids)
    cn_num = max(int(c[1:]) for c in cids)
    today = date.today().isoformat()
    note = f'KPI{month_sheet}シートから自動登録'
    mrow, lrow = wm.max_row + 1, wl.max_row + 1
    for r in missing:
        pn_num += 1
        cn_num += 1
        pid, cid = f'P{pn_num:06d}', f'C{cn_num:06d}'
        ptype = 'セット' if 'セット' in r['name'] else '単品'
        vals_m = [pid, extract_jan(r['pn']) or None, r['name'], ptype, r['cost'],
                  '楽天', '販売中', today, today, note]
        for col, v in enumerate(vals_m, start=1):
            wm.cell(mrow, col).value = v
        vals_l = [cid, pid, '楽天', r['ctrl'], r['sku'] or r['ctrl'], None, None,
                  None, '販売中', today, note]
        for col, v in enumerate(vals_l, start=1):
            wl.cell(lrow, col).value = v
        mrow += 1
        lrow += 1
    wb.save(MASTER_FILE)
    return pn_num, cn_num


def adopt_kpi_costs(month_sheet, conflict):
    """不一致分のマスター標準原価をKPI側の値で更新する(ユーザー承認済みの場合のみ)"""
    wb = load_workbook(MASTER_FILE)
    wm = wb[SH_MASTER]
    target = {pid: r['cost'] for r, pid, _ in conflict}
    today = date.today().isoformat()
    updated = 0
    for row in range(2, wm.max_row + 1):
        pid = str(wm.cell(row, PM_ID).value or '')
        if pid in target:
            wm.cell(row, PM_COST).value = target[pid]
            wm.cell(row, PM_UPD).value = today
            old_note = wm.cell(row, PM_NOTE).value
            note = f'原価をKPI{month_sheet}に合わせて更新'
            wm.cell(row, PM_NOTE).value = f'{old_note} / {note}' if old_note else note
            updated += 1
    wb.save(MASTER_FILE)
    return updated


TOOL_FILE = BASE + '/01_InventoryManagement/SourceData/楽天在庫金額集計ツール_v1.0.xlsm'
SH_TOOL_COST = '原価マスター'
AMZ_KPI_FILE = BASE + '/02_Analytics/SourceData/Amazon運営 KPI管理シート.xlsx'
AMZ_INV_FILE = BASE + '/01_InventoryManagement/SourceData/Amazon在庫リスト_import.xlsx'


def load_master_amazon():
    """マスターを読み、ASIN/AmazonSKU→内部管理ID と ID→標準原価 を返す"""
    wb = load_workbook(MASTER_FILE, data_only=True)
    cost_by_pid = {}
    for r in wb[SH_MASTER].iter_rows(min_row=2, values_only=True):
        if r[0]:
            cost_by_pid[str(r[0])] = r[PM_COST - 1]
    asin2pid, asku2pid = {}, {}
    for r in wb[SH_LISTING].iter_rows(min_row=2, values_only=True):
        pid, asin, asku = r[LT_PID - 1], r[LT_ASIN - 1], r[LT_ASKU - 1]
        if not pid:
            continue
        if asin:
            asin2pid.setdefault(norm(asin), str(pid))
        if asku:
            asku2pid.setdefault(norm(asku), str(pid))
    return cost_by_pid, asin2pid, asku2pid


def load_amazon_kpi_month(month_sheet):
    """Amazon KPI月次シートから (ASIN, AmazonSKU, 商品名, 仕入値) をペア重複なしで返す"""
    wb = load_workbook(AMZ_KPI_FILE, data_only=True)
    ws = wb[month_sheet]
    out, seen = [], set()
    for row in range(8, ws.max_row + 1):
        asin = ws.cell(row, 3).value
        if asin is None:
            break
        key = (norm(asin), norm(ws.cell(row, 2).value))
        if key in seen:
            continue
        seen.add(key)
        out.append({
            'asin': str(asin).strip(),
            'sku': str(ws.cell(row, 2).value or '').strip(),
            'name': str(ws.cell(row, 1).value or '').strip(),
            'cost': ws.cell(row, 12).value,
        })
    return out


def classify_amazon(rows, cost_by_pid, asin2pid, asku2pid):
    match, conflict, missing = [], [], []
    for r in rows:
        pid = asin2pid.get(norm(r['asin'])) or asku2pid.get(norm(r['sku']))
        if pid is None:
            missing.append(r)
            continue
        mc = cost_by_pid.get(pid)
        if mc is not None and r['cost'] is not None and float(mc) == float(r['cost']):
            match.append(r)
        else:
            conflict.append((r, pid, mc))
    return match, conflict, missing


def load_amazon_jan_lookup():
    """Amazon在庫リストから ASIN→JAN を引く(マスター登録時のJAN補完用)"""
    wb = load_workbook(AMZ_INV_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    jan_by_asin = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        jan, asin = r[0], r[1]
        if asin and jan:
            jan_by_asin.setdefault(norm(asin), str(jan).strip())
    return jan_by_asin


def register_missing_amazon(month_sheet, missing):
    """Amazon側の未登録商品を商品マスター+出品テーブルへ追加登録する"""
    jan_by_asin = load_amazon_jan_lookup()
    wb = load_workbook(MASTER_FILE)
    wm, wl = wb[SH_MASTER], wb[SH_LISTING]
    pn_num = max(int(str(r[0])[1:]) for r in wm.iter_rows(min_row=2, values_only=True) if r[0])
    cn_num = max(int(str(r[0])[1:]) for r in wl.iter_rows(min_row=2, values_only=True) if r[0])
    today = date.today().isoformat()
    note = f'Amazon KPI{month_sheet}シートから自動登録'
    mrow, lrow = wm.max_row + 1, wl.max_row + 1
    for r in missing:
        pn_num += 1
        cn_num += 1
        pid, cid = f'P{pn_num:06d}', f'C{cn_num:06d}'
        ptype = 'セット' if 'セット' in r['name'] else '単品'
        vals_m = [pid, jan_by_asin.get(norm(r['asin'])), r['name'], ptype, r['cost'],
                  'Amazon', '販売中', today, today, note]
        for col, v in enumerate(vals_m, start=1):
            wm.cell(mrow, col).value = v
        vals_l = [cid, pid, 'Amazon', None, None, r['asin'], r['sku'] or None,
                  None, '販売中', today, note]
        for col, v in enumerate(vals_l, start=1):
            wl.cell(lrow, col).value = v
        mrow += 1
        lrow += 1
    wb.save(MASTER_FILE)
    return pn_num, cn_num


def build_push_rows():
    """商品マスター+出品テーブルから集計ツール原価マスター用の行を作る
    (JAN, 楽天商品管理番号, 商品名, 商品種別, 標準原価, 最終更新日, 備考)"""
    wb = load_workbook(MASTER_FILE, data_only=True)
    products = {}
    for r in wb[SH_MASTER].iter_rows(min_row=2, values_only=True):
        if r[0]:
            products[str(r[0])] = r
    rctrl_by_pid = {}
    for r in wb[SH_LISTING].iter_rows(min_row=2, values_only=True):
        pid, ch, rctrl = r[LT_PID - 1], r[LT_CH - 1], r[LT_RCTRL - 1]
        if pid and rctrl and ch and '楽天' in str(ch):
            rctrl_by_pid.setdefault(str(pid), str(rctrl).strip())
    rows, seen = [], set()
    for pid, p in products.items():
        jan = str(p[PM_JAN - 1]).strip() if p[PM_JAN - 1] else ''
        rctrl = rctrl_by_pid.get(pid, '')
        cost = p[PM_COST - 1]
        if cost is None or (jan == '' and rctrl == ''):
            continue
        key = (jan, rctrl)
        if key in seen:
            continue
        seen.add(key)
        upd = p[PM_UPD - 1]
        upd = upd.isoformat()[:10] if hasattr(upd, 'isoformat') else (str(upd)[:10] if upd else '')
        rows.append([jan, rctrl, str(p[PM_NAME - 1] or ''), str(p[PM_TYPE - 1] or '単品'),
                     float(cost), upd, '商品マスターから同期'])
    return rows


def push_to_tool(rows):
    """Excel(AppleScript)経由で集計ツールの原価マスターシートへ書き込む"""
    import subprocess
    import tempfile

    def esc(v):
        if isinstance(v, float):
            return str(int(v)) if v == int(v) else str(v)
        return '"' + str(v).replace('\\', '\\\\').replace('"', '\\"') + '"'

    # 1行が長くなりすぎないよう100行ずつ set value する(AppleScriptの複数行リストは書けないため)
    CHUNK = 100
    stmts = []
    for i in range(0, len(rows), CHUNK):
        chunk = rows[i:i + CHUNK]
        data = ','.join('{' + ','.join(esc(v) for v in row) + '}' for row in chunk)
        r1, r2 = 3 + i, 2 + i + len(chunk)
        stmts.append(f'set value of range ("A{r1}:G{r2}") of ws to {{{data}}}')
    body = '\n    '.join(stmts)
    # 集計シートが10万行あるため、書き込み中は手動計算にして最後に一括再計算する
    script = f'''
set p to POSIX file "{TOOL_FILE}"
with timeout of 900 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    set wb to active workbook
    set ws to worksheet "{SH_TOOL_COST}" of wb
    set calculation to calculation manual
    set screen updating to false
    clear contents of range "A3:G10000" of ws
    {body}
    set calculation to calculation automatic
    calculate
    set screen updating to true
    save wb
    close wb saving no
    if not wasRunning then quit
end tell
end timeout
return "push done: {len(rows)} rows"
'''
    with tempfile.NamedTemporaryFile('w', suffix='.applescript', delete=False) as f:
        f.write(script)
        path = f.name
    r = subprocess.run(['osascript', path], capture_output=True, text=True, timeout=960)
    if r.returncode != 0:
        raise RuntimeError(f'AppleScript failed: {r.stderr}')
    return r.stdout.strip()


def main_amazon(mode, month_sheet):
    cost_by_pid, asin2pid, asku2pid = load_master_amazon()
    rows = load_amazon_kpi_month(month_sheet)
    match, conflict, missing = classify_amazon(rows, cost_by_pid, asin2pid, asku2pid)
    print(f'Amazon {month_sheet}: ユニーク{len(rows)}ペア')
    print(f'  マスター一致: {len(match)} / 原価不一致: {len(conflict)} / 未登録: {len(missing)}')
    if conflict:
        if mode == 'adopt-amazon':
            print('--- 原価不一致(これからKPI側の値でマスターを更新します)')
        else:
            print('--- 原価不一致(マスターは変更していません。要ユーザー判断)')
        for r, pid, mc in conflict:
            print(f'  {r["asin"]} ({pid}) マスター:{mc} / KPI:{r["cost"]} — {r["name"][:30]}')
    if mode == 'adopt-amazon':
        if not conflict:
            print('不一致なし — 更新処理スキップ')
            return
        updated = adopt_kpi_costs(f'Amazon {month_sheet}', conflict)
        print(f'マスター原価をKPI側に更新: {updated}件')
    elif mode == 'register-amazon':
        if not missing:
            print('未登録なし — 登録処理スキップ')
            return
        last_p, last_c = register_missing_amazon(month_sheet, missing)
        print(f'登録完了: {len(missing)}件追加 (最終ID P{last_p:06d} / C{last_c:06d})')
    elif missing:
        print(f'--- 未登録 {len(missing)}件(register-amazon モードで追加登録できます)')


def main():
    if len(sys.argv) >= 2 and sys.argv[1] == 'push':
        rows = build_push_rows()
        print(f'配信対象: {len(rows)}行(商品マスター由来)')
        print(push_to_tool(rows))
        return
    if len(sys.argv) >= 3 and sys.argv[1] in ('check-amazon', 'register-amazon', 'adopt-amazon'):
        main_amazon(sys.argv[1], sys.argv[2])
        return
    if len(sys.argv) < 3 or sys.argv[1] not in ('check', 'register', 'adopt'):
        print(__doc__)
        sys.exit(1)
    mode, month_sheet = sys.argv[1], sys.argv[2]
    cost_by_pid, pair2pid, ctrl2pid = load_master()
    rows = load_kpi_month(month_sheet)
    match, conflict, missing = classify(rows, cost_by_pid, pair2pid, ctrl2pid)
    print(f'{month_sheet}: {len(rows)}行(ユニーク{len(match)+len(conflict)+len(missing)}ペア)')
    print(f'  マスター一致: {len(match)} / 原価不一致: {len(conflict)} / 未登録: {len(missing)}')
    if conflict:
        if mode == 'adopt':
            print('--- 原価不一致(これからKPI側の値でマスターを更新します)')
        else:
            print('--- 原価不一致(マスターは変更していません。要ユーザー判断)')
        for r, pid, mc in conflict:
            print(f'  {r["ctrl"]} ({pid}) マスター:{mc} / KPI:{r["cost"]} — {r["name"][:30]}')
    if mode == 'adopt':
        if not conflict:
            print('不一致なし — 更新処理スキップ')
            return
        updated = adopt_kpi_costs(month_sheet, conflict)
        print(f'マスター原価をKPI側に更新: {updated}件')
        return
    if mode == 'register':
        if not missing:
            print('未登録なし — 登録処理スキップ')
            return
        last_p, last_c = register_missing(month_sheet, missing)
        print(f'登録完了: {len(missing)}件追加 (最終ID P{last_p:06d} / C{last_c:06d})')
    elif missing:
        print(f'--- 未登録 {len(missing)}件(register モードで追加登録できます)')


if __name__ == '__main__':
    main()
