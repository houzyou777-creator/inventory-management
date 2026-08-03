# -*- coding: utf-8 -*-
"""apply_amazon_fees.py — トランザクションレポートの実額手数料をAmazon KPIシートへ反映

使い方:
    python3 apply_amazon_fees.py <Transaction.csv> <シート名>

例:
    python3 apply_amazon_fees.py ../SourceData/2026JulMonthlyTransaction.csv 7月

処理内容:
1. トランザクションレポート(注文+返金)をSKU別に集計
   手数料 = 販売手数料 + FBA手数料 + トランザクションに関するその他の手数料
2. KPIシートのK列(販売手数料)を、SKUが一致した行だけ実額に置き換える
   一致しないSKU(決済が翌月にずれた注文など)は暫定15%の数式のまま残す
3. 経費ブロックへ自動入力: プロモーション費(プロモ割引+ポイント費用)、
   その他手数料(FBA在庫保管・返送・月額登録料などの注文外費用)、送料(購入配送ラベル実費)
   ※広告費・梱包資材はレポートに含まれないため黄色のまま
4. Excel(AppleScript)で再計算して検証

注意: ビジネスレポート(注文日基準・税込)とトランザクション(決済日基準・税抜)は
集計基準が異なるため、手数料は「その月に決済された実額」として扱う。
"""
import csv
import os
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE = '/Users/hide0726/Desktop/Claude Code/MomijiStore_OS'
OUT_FILE = os.environ.get(
    'AMZ_KPI_FILE_OVERRIDE',
    BASE + '/02_Analytics/SourceData/Amazon運営 KPI管理シート.xlsx')
NOFILL = PatternFill(fill_type=None)


def num(s):
    s = s.replace(',', '').strip()
    return float(s) if s else 0.0


def read_transactions(path):
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    hi = next(i for i, r in enumerate(rows) if r and r[0].startswith('日付/時間'))
    fee_by_sku = defaultdict(float)
    promo = ship = other = 0.0
    for r in rows[hi + 1:]:
        if len(r) < 28:
            continue
        typ = r[2]
        if typ in ('注文', '返金'):
            fee_by_sku[r[4].strip().upper()] += num(r[23]) + num(r[24]) + num(r[25])
            promo += num(r[20]) + num(r[19])
        elif typ == '配送サービス':
            ship += num(r[27])
        elif typ == '振込み':
            continue
        else:
            other += num(r[27])
    return fee_by_sku, -promo, -ship, -other


def apply(sheet_name, fee_by_sku, promo, ship, other):
    wb = load_workbook(OUT_FILE)
    ws = wb[sheet_name]

    # 同一SKUが複数行に分かれている場合があるため、まず行を集めてから売上比で按分する
    rows_by_sku = {}
    row = 8
    while ws.cell(row, 3).value is not None:     # C列(ASIN)が空になるまでがデータ
        sku = str(ws.cell(row, 2).value or '').strip().upper()
        sales = ws.cell(row, 9).value or 0
        rows_by_sku.setdefault(sku, []).append((row, sales))
        row += 1

    matched = unmatched = 0
    matched_fee = 0.0
    for sku, rlist in rows_by_sku.items():
        fee = fee_by_sku.get(sku)
        if fee is None:
            unmatched += len(rlist)               # 暫定15%の数式のまま
            continue
        total_fee = round(-fee, 2)
        total_sales = sum(s for _, s in rlist)
        assigned = 0.0
        for i, (r, s) in enumerate(rlist):
            if i == len(rlist) - 1:
                v = round(total_fee - assigned, 2)   # 端数は最終行で調整
            else:
                share = s / total_sales if total_sales else 1 / len(rlist)
                v = round(total_fee * share, 2)
                assigned += v
            ws.cell(r, 11).value = int(v) if v == int(v) else v
            matched += 1
        matched_fee += total_fee

    # 経費ブロックはM列のラベルで行を特定する(行番号のハードコード回避)
    label_row = {}
    for r in range(row, ws.max_row + 1):
        lab = ws.cell(r, 13).value
        if lab:
            label_row.setdefault(str(lab), r)
    for lab, val in [('プロモーション費', promo), ('その他手数料', other), ('送料', ship)]:
        r = label_row.get(lab)
        if r is None:
            raise SystemExit(f'経費ラベル「{lab}」が見つかりません')
        c = ws.cell(r, 14)
        c.value = round(val)
        c.fill = NOFILL

    # ヘッダーの注記を実額版に更新
    ws.cell(5, 2).value = '実額(トランザクションレポートより)。未マッチSKUのみ暫定15%'

    wb.save(OUT_FILE)
    return matched, unmatched, matched_fee


def recalc_via_excel(path):
    script = f'''
set p to POSIX file "{path}"
with timeout of 600 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    calculate
    save active workbook
    close active workbook saving no
    if not wasRunning then quit
end tell
end timeout
return "ok"
'''
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=660)
    if r.returncode != 0:
        raise RuntimeError(f'Excel再計算に失敗: {r.stderr}')


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    csv_path, sheet_name = sys.argv[1], sys.argv[2]

    bdir = os.path.dirname(OUT_FILE) + '/Backup'
    os.makedirs(bdir, exist_ok=True)
    base = os.path.splitext(os.path.basename(OUT_FILE))[0]
    shutil.copy2(OUT_FILE, f'{bdir}/{base}_backup_{date.today():%Y%m%d}_{sheet_name}手数料実額化前.xlsx')
    print('バックアップ取得済み')

    fee_by_sku, promo, ship, other = read_transactions(csv_path)
    print(f'トランザクション集計: SKU {len(fee_by_sku)}件 / プロモ費 ¥{promo:,.0f} / 送料 ¥{ship:,.0f} / その他 ¥{other:,.0f}')

    matched, unmatched, matched_fee = apply(sheet_name, fee_by_sku, promo, ship, other)
    print(f'K列置換: 実額 {matched}行 (¥{matched_fee:,.0f}) / 暫定15%のまま {unmatched}行')

    print('Excelで再計算中...')
    recalc_via_excel(OUT_FILE)

    wb = load_workbook(OUT_FILE, data_only=True)
    ws = wb[sheet_name]
    errs = [c.coordinate for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('#')]
    t = 8
    while ws.cell(t, 3).value is not None:
        t += 1
    print(f'数式エラー: {len(errs)}件')
    print(f'手数料合計 K{t}: ¥{ws.cell(t, 11).value:,.0f} / 粗利 N{t}: ¥{ws.cell(t, 14).value:,.0f} ({ws.cell(t, 15).value * 100:.1f}%)')
    gr = next((r for r in range(t, ws.max_row + 1) if ws.cell(r, 13).value == '限界利益'), None)
    if gr:
        v = ws.cell(gr, 14).value
        vr = ws.cell(gr + 1, 14).value
        print(f'限界利益: ¥{v:,.0f} ({vr * 100:.1f}%)' if v is not None else '')
    if errs:
        sys.exit('*** FAIL ***')
    print('PASS')


if __name__ == '__main__':
    main()
