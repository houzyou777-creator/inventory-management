# -*- coding: utf-8 -*-
"""build_rakuten_rpp_month.py — 楽天RPP広告レポートから月次広告分析シートを生成

使い方:
    python3 build_rakuten_rpp_month.py <rpp_item_reports.csv> <rpp_reports.csv> <シート名>

例:
    python3 build_rakuten_rpp_month.py ../SourceData/rpp_item_reports_xxx.csv \\
        ../SourceData/rpp_reports_xxx.csv 7月

処理内容:
1. 出力先「楽天RPP広告分析.xlsx」に月別シートを生成(既存ならBackup/へバックアップ)
2. 商品別レポート(cp932)を広告費降順で展開し、KPI管理シートの同月シートと
   商品管理番号で結合して「広告依存度」「広告費対粗利」を算出
3. 売上・ROASは720時間(30日)アトリビューションを主指標、12時間を参考列にする
4. Excel(AppleScript)で再計算し、広告費合計をサマリーレポートと突合して検証

レポートはRMS広告センター(プロモーションメニュー → RPP)からダウンロードする。
"""
import csv
import os
import shutil
import subprocess
import sys
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

BASE = '/Users/hide0726/Desktop/Claude Code/MomijiStore_OS'
OUT_FILE = os.environ.get(
    'RPP_FILE_OVERRIDE',
    BASE + '/02_Analytics/SourceData/楽天RPP広告分析.xlsx')
KPI_FILE = BASE + '/02_Analytics/SourceData/楽天運営 KPI管理シート.xlsx'

FONT = Font(name='Arial', size=10)
FONT_B = Font(name='Arial', size=10, bold=True)

HEADERS = ['商品管理番号', '商品名', '入札単価', 'クリック数', 'CTR', '広告費(実績額)',
           'CPC実績', '広告経由売上(720h)', '売上件数(720h)', 'CVR(720h)', 'ROAS(720h)',
           '広告経由売上(12h)', '商品月間売上', '広告依存度', '商品月間粗利', '広告費/粗利']


def fnum(s, as_int=True):
    s = str(s).replace(',', '').replace('%', '').strip()
    if s == '':
        return None
    v = float(s)
    return int(v) if as_int and v == int(v) else v


def read_item_report(path):
    with open(path, encoding='cp932') as f:
        rows = list(csv.reader(f))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == 'コントロールカラム')
    data = []
    for r in rows[hi + 1:]:
        if len(r) < 46 or not r[3].strip():
            continue
        data.append({
            'ctrl': r[3].strip(), 'bid': fnum(r[4]), 'ctr': fnum(r[5], False),
            'clicks': fnum(r[7]), 'spend': fnum(r[8]), 'cpc': fnum(r[9]),
            'sales720': fnum(r[21]), 'orders720': fnum(r[22]), 'cvr720': fnum(r[23], False),
            'sales12': fnum(r[16]),
        })
    data.sort(key=lambda d: -(d['spend'] or 0))
    return data


def read_summary(path):
    with open(path, encoding='cp932') as f:
        rows = list(csv.reader(f))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == '日付')
    h, d = rows[hi], rows[hi + 1]
    return dict(zip(h, d))


def load_kpi_month(month_sheet):
    """KPI月次シートから 管理番号→(商品名, 売上合計, 粗利合計) を作る"""
    wb = load_workbook(KPI_FILE, data_only=True)
    if month_sheet not in wb.sheetnames:
        return {}
    ws = wb[month_sheet]
    agg = {}
    for row in range(8, ws.max_row + 1):
        ctrl = ws.cell(row, 3).value
        if ctrl is None:
            break
        key = str(ctrl).strip().lower()
        name = str(ws.cell(row, 1).value or '')
        sales = ws.cell(row, 9).value or 0
        profit = ws.cell(row, 14).value or 0
        if key in agg:
            agg[key][1] += sales
            agg[key][2] += profit
        else:
            agg[key] = [name, sales, profit]
    return agg


def build_sheet(items, summary, month_sheet, kpi):
    if os.path.exists(OUT_FILE):
        wb = load_workbook(OUT_FILE)
        if month_sheet in wb.sheetnames:
            raise SystemExit(f'シート「{month_sheet}」は既に存在します。削除してから再実行してください。')
        ws = wb.create_sheet(month_sheet)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = month_sheet

    def put(row, col, val, font=FONT, fmt=None):
        c = ws.cell(row, col)
        c.value = val
        c.font = font
        if fmt:
            c.number_format = fmt
        return c

    put(1, 1, 'RPP広告分析', FONT_B)
    put(2, 1, '集計期間')
    put(2, 2, summary.get('日付', month_sheet))
    labels = [('有効予算', fnum(summary.get('有効予算'))),
              ('消化率', fnum(summary.get('消化率(%)'), False) / 100),
              ('広告費(割引後実績額)', fnum(summary.get('割引後実績額'))),
              ('クリック数', fnum(summary.get('クリック数(合計)'))),
              ('CPC実績', fnum(summary.get('CPC実績(合計)'))),
              ('広告経由売上(720h)', fnum(summary.get('売上金額(合計720時間)'))),
              ('ROAS(720h)', fnum(summary.get('ROAS(合計720時間)(%)'), False) / 100),
              ('CVR(720h)', fnum(summary.get('CVR(合計720時間)(%)'), False) / 100)]
    for j, (lab, v) in enumerate(labels):
        put(3 + j, 1, lab, FONT_B)
        fmt = '0.0%' if lab in ('消化率', 'ROAS(720h)', 'CVR(720h)') else '#,##0'
        put(3 + j, 2, v, fmt=fmt)

    HR = 12                                  # 見出し行
    for col, h in enumerate(HEADERS, start=1):
        put(HR, col, h, FONT_B)

    r0 = HR + 1
    for i, d in enumerate(items):
        row = r0 + i
        k = kpi.get(d['ctrl'].lower())
        put(row, 1, d['ctrl'])
        put(row, 2, (k[0][:60] if k else ''))
        put(row, 3, d['bid'])
        put(row, 4, d['clicks'])
        put(row, 5, (d['ctr'] or 0) / 100, fmt='0.00%')
        put(row, 6, d['spend'])
        put(row, 7, d['cpc'])
        put(row, 8, d['sales720'])
        put(row, 9, d['orders720'])
        put(row, 10, (d['cvr720'] or 0) / 100, fmt='0.00%')
        put(row, 11, f'=IF(F{row}=0,"",H{row}/F{row})', fmt='0%')
        put(row, 12, d['sales12'])
        put(row, 13, k[1] if k else None, fmt='#,##0')
        put(row, 14, f'=IF(OR(M{row}="",M{row}=0),"",H{row}/M{row})', fmt='0.0%')
        put(row, 15, round(k[2]) if k else None, fmt='#,##0')
        put(row, 16, f'=IF(OR(O{row}="",O{row}=0),"",F{row}/O{row})', fmt='0.0%')

    last = r0 + len(items) - 1
    t = last + 1
    put(t, 1, '合計', FONT_B)
    for col_l in 'DFHIL':
        col = 'ABCDEFGHIJKLMNOP'.index(col_l) + 1
        put(t, col, f'=SUM({col_l}{r0}:{col_l}{last})', FONT_B)
    put(t, 11, f'=IF(F{t}=0,"",H{t}/F{t})', FONT_B, fmt='0%')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    for c in 'FHLMO':
        ws.column_dimensions[c].width = 13
    wb.save(OUT_FILE)
    return t


def recalc_via_excel(path):
    script = f'''
set p to POSIX file "{path}"
with timeout of 600 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    calculate
    save active workbook
    close active workbook saving no
    if not wasRunning then quit
end tell
end timeout
return "ok"
'''
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=660)
    if r.returncode != 0:
        raise RuntimeError(f'Excel再計算に失敗: {r.stderr}')


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    item_csv, summary_csv, month_sheet = sys.argv[1], sys.argv[2], sys.argv[3]

    if os.path.exists(OUT_FILE):
        bdir = os.path.dirname(OUT_FILE) + '/Backup'
        os.makedirs(bdir, exist_ok=True)
        base = os.path.splitext(os.path.basename(OUT_FILE))[0]
        shutil.copy2(OUT_FILE, f'{bdir}/{base}_backup_{date.today():%Y%m%d}_{month_sheet}生成前.xlsx')
        print('バックアップ取得済み')

    items = read_item_report(item_csv)
    summary = read_summary(summary_csv)
    kpi = load_kpi_month(month_sheet)
    print(f'{month_sheet}: 商品 {len(items)}件 / KPIシート結合 {sum(1 for d in items if d["ctrl"].lower() in kpi)}件')

    t = build_sheet(items, summary, month_sheet, kpi)
    print('Excelで再計算中...')
    recalc_via_excel(OUT_FILE)

    wb = load_workbook(OUT_FILE, data_only=True)
    ws = wb[month_sheet]
    errs = [c.coordinate for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('#')]
    spend_sum = ws.cell(t, 6).value
    spend_rep = fnum(summary.get('実績額(合計)'))
    ok = spend_sum == spend_rep
    print(f'検証: 広告費合計 {"一致" if ok else "不一致!"} (商品別合計 {spend_sum:,} / サマリー {spend_rep:,})')
    print(f'数式エラー: {len(errs)}件')
    if not ok or errs:
        sys.exit('*** FAIL ***')
    print('PASS')


if __name__ == '__main__':
    main()
