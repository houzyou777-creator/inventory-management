# -*- coding: utf-8 -*-
"""build_amazon_kpi_month.py — Amazonビジネスレポート(子ASIN別)から月次KPIシートを生成

使い方:
    python3 build_amazon_kpi_month.py <BusinessReport.csv> <シート名>

例:
    python3 build_amazon_kpi_month.py ../SourceData/BusinessReport-03-08-26.csv 7月

処理内容:
1. 出力先「Amazon運営 KPI管理シート.xlsx」が既にあればBackup/へバックアップ
2. 売上降順で商品を並べ、楽天KPIシートと同じK〜O列(手数料・仕入値・粗利)を付与
   + Amazon固有のセッション数・ユニットセッション率列
3. 仕入値の照合順: 商品マスター(ASIN→AmazonSKU) → Amazon在庫リスト(ASIN→SKU管理番号)
   → 解決不能は黄色塗り空欄
4. 販売手数料は暫定一律15%(トランザクションレポート入手後に実額へ置き換える前提。
   暫定である旨をシート上部に明記する)
5. Excel(AppleScript)で再計算し、合計をCSVと突合して検証

月次経費(広告費・FBA手数料・送料等)は黄色塗り空欄 — ユーザー入力後に限界利益が確定。
"""
import csv
import os
import re
import shutil
import subprocess
import sys
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

BASE = '/Users/hide0726/Desktop/Claude Code/MomijiStore_OS'
OUT_FILE = os.environ.get(
    'AMZ_KPI_FILE_OVERRIDE',
    BASE + '/02_Analytics/SourceData/Amazon運営 KPI管理シート.xlsx')
MASTER_FILE = BASE + '/01_InventoryManagement/SourceData/商品マスター_単品_v1.0.xlsx'
AMZ_INV_FILE = BASE + '/01_InventoryManagement/SourceData/Amazon在庫リスト_import.xlsx'

YELLOW = PatternFill('solid', fgColor='FFFF00')
FONT = Font(name='Arial', size=10)
FONT_B = Font(name='Arial', size=10, bold=True)

HEADERS = ['商品名', 'SKU', 'ASIN', '親ASIN', 'セッション数', 'ユニットセッション率',
           '平均単価', '売上個数', '売上', '売上件数', '販売手数料', '仕入値',
           '仕入値合計', '利益(粗利)', '利益率（粗利率）']


def yen(s):
    return int(s.replace('￥', '').replace(',', '').strip() or 0)


def num(s):
    return int(s.replace(',', '').strip() or 0)


def pct(s):
    s = s.replace('%', '').strip()
    return float(s) / 100 if s else None


def read_report(path):
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    data = []
    for r in rows[1:]:
        if len(r) < 21 or not any(x.strip() for x in r):
            continue
        data.append({
            'parent': r[0].strip(), 'asin': r[1].strip(), 'name': r[2].strip(),
            'sku': r[3].strip(), 'sessions': num(r[4]), 'usr': pct(r[16]),
            'units': num(r[14]), 'sales': yen(r[18]), 'orders': num(r[20]),
        })
    data.sort(key=lambda d: -d['sales'])
    return data


def build_cost_resolver():
    """商品マスター → Amazon在庫リスト の順で原価を引く関数を返す"""
    wbm = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    cost_by_pid = {str(r[0]): r[4] for r in wbm['商品マスター'].iter_rows(min_row=2, values_only=True) if r[0]}
    asin2pid, asku2pid = {}, {}
    for r in wbm['出品テーブル'].iter_rows(min_row=2, values_only=True):
        pid, asin, asku = r[1], r[5], r[6]
        if not pid:
            continue
        if asin:
            asin2pid.setdefault(str(asin).strip().upper(), str(pid))
        if asku:
            asku2pid.setdefault(str(asku).strip().upper(), str(pid))
    wba = load_workbook(AMZ_INV_FILE, read_only=True, data_only=True)
    inv_asin, inv_sku = {}, {}
    for r in wba[wba.sheetnames[0]].iter_rows(min_row=2, values_only=True):
        asin, skum, cost = r[1], r[3], r[6]
        if cost is None:
            continue
        try:
            cost = float(cost)
        except (TypeError, ValueError):
            continue
        if asin:
            inv_asin.setdefault(str(asin).strip().upper(), cost)
        if skum:
            inv_sku.setdefault(str(skum).strip().upper(), cost)

    def resolve(asin, sku):
        a, s = asin.upper(), sku.upper()
        pid = asin2pid.get(a) or asku2pid.get(s)
        if pid is not None and cost_by_pid.get(pid) is not None:
            return cost_by_pid[pid]
        return inv_asin.get(a) or inv_sku.get(s)

    return resolve


def build_sheet(data, sheet_name, source_name, resolve):
    if os.path.exists(OUT_FILE):
        wb = load_workbook(OUT_FILE)
        if sheet_name in wb.sheetnames:
            raise SystemExit(f'シート「{sheet_name}」は既に存在します。削除してから再実行してください。')
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    def put(row, col, val, font=FONT, fmt=None, fill=None):
        c = ws.cell(row, col)
        c.value = val
        c.font = font
        if fmt:
            c.number_format = fmt
        if fill is not None:
            c.fill = fill
        return c

    put(1, 1, '※この情報は店舗様およびAmazonでの重要な情報となります。データの取扱には十分にご注意ください。')
    put(2, 1, 'ASIN別売上(ビジネスレポート)')
    put(3, 1, '表示期間')
    put(3, 2, sheet_name)
    put(4, 1, '出典')
    put(4, 2, source_name)
    put(5, 1, '販売手数料')
    put(5, 2, '暫定一律15%(トランザクションレポート入手後に実額へ置換)')
    for col, h in enumerate(HEADERS, start=1):
        put(7, col, h)

    n = len(data)
    missing = []
    for i, d in enumerate(data):
        row = 8 + i
        put(row, 1, d['name'])
        put(row, 2, d['sku'])
        put(row, 3, d['asin'])
        put(row, 4, d['parent'])
        put(row, 5, d['sessions'])
        put(row, 6, d['usr'], fmt='0.00%')
        put(row, 7, f'=IF(H{row}=0,"",I{row}/H{row})', fmt='0')
        put(row, 8, d['units'])
        put(row, 9, d['sales'])
        put(row, 10, d['orders'])
        put(row, 11, f'=SUM(I{row}*0.15)')
        cost = resolve(d['asin'], d['sku'])
        cost = int(cost) if isinstance(cost, float) and cost == int(cost) else cost
        put(row, 12, cost, fill=None if cost is not None else YELLOW)
        if cost is None:
            missing.append((row, d['asin'], d['sku'], d['sales']))
        put(row, 13, f'=SUM(H{row}*L{row})')
        put(row, 14, f'=SUM(I{row}-K{row}-M{row})')
        put(row, 15, f'=IF(I{row}=0,"",N{row}/I{row})', fmt='0.0%')

    last = 7 + n
    t = last + 1
    for col_l in 'EHIJKMN':
        col = 'ABCDEFGHIJKLMNO'.index(col_l) + 1
        put(t, col, f'=SUM({col_l}8:{col_l}{last})')
    put(t, 15, f'=IF(I{t}=0,"",N{t}/I{t})', fmt='0.0%')

    e0 = t + 2
    for j, lab in enumerate(['広告費', 'プロモーション費', 'その他手数料']):
        row = e0 + j
        put(row, 13, lab, font=FONT_B)
        put(row, 14, None, fill=YELLOW)
        put(row, 15, f'=SUM(N{row}/I{t})', fmt='0.00%')
    tot1 = e0 + 3
    put(tot1, 13, '合計', font=FONT_B)
    put(tot1, 14, f'=SUM(N{e0}:N{tot1 - 1})')
    put(tot1, 15, f'=SUM(N{tot1}/I{t})', fmt='0.00%')

    s0 = tot1 + 2
    for j, lab in enumerate(['送料', '梱包資材']):
        row = s0 + j
        put(row, 13, lab, font=FONT_B)
        put(row, 14, None, fill=YELLOW)
        put(row, 15, f'=SUM(N{row}/I{t})', fmt='0.00%')
    tot2 = s0 + 2
    put(tot2, 13, '合計', font=FONT_B)
    put(tot2, 14, f'=SUM(N{s0}:N{s0 + 1})')
    put(tot2, 15, f'=SUM(O{s0}:O{s0 + 1})', fmt='0.00%')

    g = tot2 + 2
    put(g, 13, '限界利益', font=FONT_B)
    put(g, 14, f'=SUM(N{t}-N{tot1}-N{tot2})')
    put(g + 1, 13, '限界利益率', font=FONT_B)
    put(g + 1, 14, f'=IF(I{t}=0,"",N{g}/I{t})', fmt='0.0%')

    ws.column_dimensions['A'].width = 29
    ws.column_dimensions['B'].width = 20
    for c in 'CD':
        ws.column_dimensions[c].width = 13
    wb.save(OUT_FILE)
    return t, missing


def recalc_via_excel(path):
    script = f'''
set p to POSIX file "{path}"
with timeout of 600 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    calculate
    save active workbook
    close active workbook saving no
    if not wasRunning then quit
end tell
end timeout
return "ok"
'''
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=660)
    if r.returncode != 0:
        raise RuntimeError(f'Excel再計算に失敗: {r.stderr}')


def verify(sheet_name, data, total_row):
    wb = load_workbook(OUT_FILE, data_only=True)
    ws = wb[sheet_name]
    exp = [sum(d['units'] for d in data), sum(d['sales'] for d in data), sum(d['orders'] for d in data)]
    got = [ws.cell(total_row, c).value for c in (8, 9, 10)]
    errs = [c.coordinate for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('#')]
    return got == exp, exp, got, errs, ws.cell(total_row, 14).value, ws.cell(total_row, 15).value


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    csv_path, sheet_name = sys.argv[1], sys.argv[2]

    if os.path.exists(OUT_FILE):
        bdir = os.path.dirname(OUT_FILE) + '/Backup'
        os.makedirs(bdir, exist_ok=True)
        base = os.path.splitext(os.path.basename(OUT_FILE))[0]
        shutil.copy2(OUT_FILE, f'{bdir}/{base}_backup_{date.today():%Y%m%d}_{sheet_name}生成前.xlsx')
        print('バックアップ取得済み')

    data = read_report(csv_path)
    print(f'{sheet_name}: レポート {len(data)}行 / 売上合計 ¥{sum(d["sales"] for d in data):,}')
    resolve = build_cost_resolver()
    total_row, missing = build_sheet(data, sheet_name, os.path.basename(csv_path), resolve)
    print(f'シート生成完了(仕入値未解決 {len(missing)}行)')

    print('Excelで再計算中...')
    recalc_via_excel(OUT_FILE)

    ok, exp, got, errs, n_total, o_total = verify(sheet_name, data, total_row)
    print(f'検証: 合計{"一致" if ok else "不一致!"} (個数/売上/件数 CSV={exp} シート={got})')
    print(f'数式エラー: {len(errs)}件')
    if n_total is not None and o_total is not None:
        print(f'粗利(仕入値未解決{len(missing)}行を除く): {n_total:,.0f}円 ({o_total * 100:.1f}%)')
    if not ok or errs:
        sys.exit('*** FAIL — シートを確認してください ***')
    print('PASS。黄色セル(仕入値・経費)入力後に確定します。')


if __name__ == '__main__':
    main()
