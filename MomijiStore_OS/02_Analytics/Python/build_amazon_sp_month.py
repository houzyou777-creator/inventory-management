# -*- coding: utf-8 -*-
"""build_amazon_sp_month.py — Amazonスポンサープロダクト広告の月次分析シートを生成

使い方:
    python3 build_amazon_sp_month.py <広告対象商品レポート.xlsx> <シート名> [検索用語レポート.xlsx]

例:
    python3 build_amazon_sp_month.py "../SourceData/スポンサープロダクト広告_広告対象商品_レポート.xlsx" 7月 \\
        "../SourceData/スポンサープロダクト広告_検索用語_レポート (1).xlsx"

処理内容:
1. 出力先「AmazonSP広告分析.xlsx」に月別シートを生成(既存ならBackup/へバックアップ)
2. 広告対象商品レポートをSKU×ASINで集約し広告費降順で展開、
   Amazon KPI管理シートの同月とASIN/SKUで結合して広告依存度・広告費対粗利を算出
3. 検索用語レポートがあれば「<月>_検索用語」シートも生成(キーワード集約・費用降順)
4. Excel(AppleScript)で再計算し、費用合計をレポート原本と突合して検証

注意: Amazonのアトリビューションは広告クリック後7日間(楽天RPPの720時間=30日より短い)。
レポートは広告コンソール → 測定と報告 → スポンサー広告レポートから取得する。
"""
import os
import shutil
import subprocess
import sys
import warnings
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

BASE = '/Users/hide0726/Desktop/Claude Code/MomijiStore_OS'
OUT_FILE = os.environ.get(
    'SP_FILE_OVERRIDE',
    BASE + '/02_Analytics/SourceData/AmazonSP広告分析.xlsx')
AMZ_KPI_FILE = BASE + '/02_Analytics/SourceData/Amazon運営 KPI管理シート.xlsx'

FONT = Font(name='Arial', size=10)
FONT_B = Font(name='Arial', size=10, bold=True)

HEADERS = ['SKU', 'ASIN', '商品名', 'インプレッション', 'クリック数', 'CTR', 'CPC',
           '広告費', '広告経由売上(7日)', '注文数(7日)', 'CVR', 'ROAS',
           '商品月間売上', '広告依存度', '商品月間粗利', '広告費/粗利']


def read_item_report(path):
    """広告対象商品レポートを (SKU, ASIN) で集約して返す"""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    agg = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        sku, asin = r[7], r[8]
        if not asin:
            continue
        key = (str(sku or '').strip(), str(asin).strip())
        a = agg.setdefault(key, {'imp': 0, 'clicks': 0, 'spend': 0.0, 'sales': 0.0, 'orders': 0})
        a['imp'] += r[9] or 0
        a['clicks'] += r[10] or 0
        a['spend'] += r[13] or 0
        a['sales'] += r[14] or 0
        a['orders'] += r[17] or 0
    items = [{'sku': k[0], 'asin': k[1], **v} for k, v in agg.items()]
    items.sort(key=lambda d: -d['spend'])
    return items


def read_search_terms(path):
    """検索用語レポートをキーワードで集約して返す"""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    agg = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        kw = str(r[9] or '').strip()
        if not kw:
            continue
        a = agg.setdefault(kw, {'imp': 0, 'clicks': 0, 'spend': 0.0, 'sales': 0.0, 'orders': 0})
        a['imp'] += r[10] or 0
        a['clicks'] += r[11] or 0
        a['spend'] += r[14] or 0
        a['sales'] += r[15] or 0
        a['orders'] += r[18] or 0
    terms = [{'kw': k, **v} for k, v in agg.items()]
    terms.sort(key=lambda d: -d['spend'])
    return terms


def load_kpi_month(month_sheet):
    """Amazon KPI月次シートから ASIN→(商品名, 売上合計, 粗利合計) を作る(SKU→も併設)"""
    wb = load_workbook(AMZ_KPI_FILE, data_only=True)
    if month_sheet not in wb.sheetnames:
        return {}, {}
    ws = wb[month_sheet]
    by_asin, by_sku = {}, {}
    for row in range(8, ws.max_row + 1):
        asin = ws.cell(row, 3).value
        if asin is None:
            break
        name = str(ws.cell(row, 1).value or '')
        sales = ws.cell(row, 9).value or 0
        profit = ws.cell(row, 14).value or 0
        ak = str(asin).strip().upper()
        if ak in by_asin:
            by_asin[ak][1] += sales
            by_asin[ak][2] += profit
        else:
            by_asin[ak] = [name, sales, profit]
        sk = str(ws.cell(row, 2).value or '').strip().upper()
        if sk:
            by_sku.setdefault(sk, by_asin[ak])
    return by_asin, by_sku


def open_out(month_sheet, suffix=''):
    name = month_sheet + suffix
    if os.path.exists(OUT_FILE):
        wb = load_workbook(OUT_FILE)
        if name in wb.sheetnames:
            raise SystemExit(f'シート「{name}」は既に存在します。削除してから再実行してください。')
        ws = wb.create_sheet(name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = name
    return wb, ws


def put(ws, row, col, val, font=FONT, fmt=None):
    c = ws.cell(row, col)
    c.value = val
    c.font = font
    if fmt:
        c.number_format = fmt
    return c


def build_item_sheet(items, month_sheet, by_asin, by_sku):
    wb, ws = open_out(month_sheet)
    put(ws, 1, 1, 'スポンサープロダクト広告分析', FONT_B)
    put(ws, 2, 1, '集計期間')
    put(ws, 2, 2, month_sheet)
    put(ws, 3, 1, 'アトリビューション')
    put(ws, 3, 2, '広告クリック後7日間(楽天RPPの720hより短い点に注意)')

    HR = 5
    for col, h in enumerate(HEADERS, start=1):
        put(ws, HR, col, h, FONT_B)
    r0 = HR + 1
    for i, d in enumerate(items):
        row = r0 + i
        k = by_asin.get(d['asin'].upper()) or by_sku.get(d['sku'].upper())
        put(ws, row, 1, d['sku'])
        put(ws, row, 2, d['asin'])
        put(ws, row, 3, (k[0][:60] if k else ''))
        put(ws, row, 4, d['imp'])
        put(ws, row, 5, d['clicks'])
        put(ws, row, 6, f'=IF(D{row}=0,"",E{row}/D{row})', fmt='0.00%')
        put(ws, row, 7, f'=IF(E{row}=0,"",H{row}/E{row})', fmt='0')
        put(ws, row, 8, round(d['spend'], 2))
        put(ws, row, 9, round(d['sales'], 2))
        put(ws, row, 10, d['orders'])
        put(ws, row, 11, f'=IF(E{row}=0,"",J{row}/E{row})', fmt='0.00%')
        put(ws, row, 12, f'=IF(H{row}=0,"",I{row}/H{row})', fmt='0%')
        put(ws, row, 13, k[1] if k else None, fmt='#,##0')
        put(ws, row, 14, f'=IF(OR(M{row}="",M{row}=0),"",I{row}/M{row})', fmt='0.0%')
        put(ws, row, 15, round(k[2]) if k else None, fmt='#,##0')
        put(ws, row, 16, f'=IF(OR(O{row}="",O{row}=0),"",H{row}/O{row})', fmt='0.0%')
    last = r0 + len(items) - 1
    t = last + 1
    put(ws, t, 1, '合計', FONT_B)
    for col_l in 'DEHIJM':
        col = 'ABCDEFGHIJKLMNOP'.index(col_l) + 1
        put(ws, t, col, f'=SUM({col_l}{r0}:{col_l}{last})', FONT_B)
    put(ws, t, 12, f'=IF(H{t}=0,"",I{t}/H{t})', FONT_B, fmt='0%')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 40
    for c in 'HIMO':
        ws.column_dimensions[c].width = 12
    wb.save(OUT_FILE)
    return t


def build_term_sheet(terms, month_sheet):
    wb, ws = open_out(month_sheet, '_検索用語')
    heads = ['検索キーワード', 'インプレッション', 'クリック数', 'CTR', 'CPC', '広告費',
             '売上(7日)', '注文数', 'CVR', 'ROAS']
    for col, h in enumerate(heads, start=1):
        put(ws, 1, col, h, FONT_B)
    r0 = 2
    for i, d in enumerate(terms):
        row = r0 + i
        put(ws, row, 1, d['kw'])
        put(ws, row, 2, d['imp'])
        put(ws, row, 3, d['clicks'])
        put(ws, row, 4, f'=IF(B{row}=0,"",C{row}/B{row})', fmt='0.00%')
        put(ws, row, 5, f'=IF(C{row}=0,"",F{row}/C{row})', fmt='0')
        put(ws, row, 6, round(d['spend'], 2))
        put(ws, row, 7, round(d['sales'], 2))
        put(ws, row, 8, d['orders'])
        put(ws, row, 9, f'=IF(C{row}=0,"",H{row}/C{row})', fmt='0.00%')
        put(ws, row, 10, f'=IF(F{row}=0,"",G{row}/F{row})', fmt='0%')
    last = r0 + len(terms) - 1
    t = last + 1
    put(ws, t, 1, '合計', FONT_B)
    for col_l in 'BCFGH':
        col = 'ABCDEFGHIJ'.index(col_l) + 1
        put(ws, t, col, f'=SUM({col_l}{r0}:{col_l}{last})', FONT_B)
    ws.column_dimensions['A'].width = 34
    wb.save(OUT_FILE)


def recalc_via_excel(path):
    script = f'''
set p to POSIX file "{path}"
with timeout of 600 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    calculate
    save active workbook
    close active workbook saving no
    if not wasRunning then quit
end tell
end timeout
return "ok"
'''
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=660)
    if r.returncode != 0:
        raise RuntimeError(f'Excel再計算に失敗: {r.stderr}')


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    item_xlsx, month_sheet = sys.argv[1], sys.argv[2]
    term_xlsx = sys.argv[3] if len(sys.argv) > 3 else None

    if os.path.exists(OUT_FILE):
        bdir = os.path.dirname(OUT_FILE) + '/Backup'
        os.makedirs(bdir, exist_ok=True)
        base = os.path.splitext(os.path.basename(OUT_FILE))[0]
        shutil.copy2(OUT_FILE, f'{bdir}/{base}_backup_{date.today():%Y%m%d}_{month_sheet}生成前.xlsx')
        print('バックアップ取得済み')

    items = read_item_report(item_xlsx)
    exp_spend = round(sum(d['spend'] for d in items), 2)
    by_asin, by_sku = load_kpi_month(month_sheet)
    joined = sum(1 for d in items if d['asin'].upper() in by_asin or d['sku'].upper() in by_sku)
    print(f'{month_sheet}: 商品 {len(items)}件 / KPIシート結合 {joined}件 / 費用合計 ¥{exp_spend:,.0f}')

    t = build_item_sheet(items, month_sheet, by_asin, by_sku)
    if term_xlsx:
        terms = read_search_terms(term_xlsx)
        build_term_sheet(terms, month_sheet)
        print(f'検索用語: {len(terms)}キーワード')

    print('Excelで再計算中...')
    recalc_via_excel(OUT_FILE)

    wb = load_workbook(OUT_FILE, data_only=True)
    ws = wb[month_sheet]
    errs = [c.coordinate for sh in wb.sheetnames for row in wb[sh].iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('#')]
    got = round(ws.cell(t, 8).value, 2)
    ok = got == exp_spend
    print(f'検証: 広告費合計 {"一致" if ok else "不一致!"} (シート {got:,.0f} / レポート {exp_spend:,.0f})')
    print(f'数式エラー: {len(errs)}件')
    if not ok or errs:
        sys.exit('*** FAIL ***')
    print('PASS')


if __name__ == '__main__':
    main()
