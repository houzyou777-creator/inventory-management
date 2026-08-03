# -*- coding: utf-8 -*-
"""build_kpi_month.py — KPI管理シートの月次シートをRMSのSKU別売上CSVから自動生成

使い方:
    python3 build_kpi_month.py <SKU_SalesList.csv> [シート名]

例:
    python3 build_kpi_month.py ../SourceData/202608_SKU_SalesList.csv
    # シート名省略時はCSVファイル名の 2026MM から「8月」を自動判定

処理内容:
1. KPI管理シートをBackup/へバックアップ
2. 直近の月シートをテンプレートとしてコピーし、CSVデータを流し込む
3. 仕入値の照合順: 商品マスター(管理番号×SKUペア) → RMS商品番号の「/仕入値」埋め込み
   → 商品番号が6桁以下の数値ならその値 → 過去月シート → 解決不能は黄色塗り空欄
4. Excel(AppleScript)で再計算し、合計値をCSVと突合して検証
5. 月次経費(広告費・送料等)は黄色塗り空欄 — ユーザー入力後に限界利益が確定

生成後は sync_cost_master.py register で新商品をマスターへ還流させること。
"""
import csv
import os
import re
import subprocess
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from sync_cost_master import KPI_FILE, load_master, norm

YELLOW = PatternFill('solid', fgColor='FFFF00')
BACKUP_DIR = os.path.dirname(KPI_FILE) + '/Backup'


def read_rms_csv(path):
    """RMSのSKU別売上CSVを (ヘッダー6行, データ行list) で返す"""
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    head6 = rows[:6]
    data = [r for r in rows[7:] if any(x.strip() for x in r)]
    return head6, data


def conv(s):
    """CSV文字列をExcel向けに数値変換(数値でなければ文字列のまま)"""
    s = s.strip()
    if s == '':
        return None
    if s.isdigit():
        return int(s)
    try:
        return float(s)
    except ValueError:
        return s


def build_cost_resolver(wb_values):
    """仕入値の解決関数を作る(商品マスター → RMS埋め込み → 過去月シート)"""
    cost_by_pid, pair2pid, ctrl2pid = load_master()

    hist_pair = {}
    for name in reversed(wb_values.sheetnames):        # 新しい月を優先
        ws = wb_values[name]
        for r in range(8, ws.max_row + 1):
            c, e, l = ws.cell(r, 3).value, ws.cell(r, 5).value, ws.cell(r, 12).value
            if c is None or not isinstance(l, (int, float)):
                continue
            hist_pair.setdefault((norm(c), norm(e)), l)

    def resolve(ctrl, pn, sku):
        key = (norm(ctrl), norm(sku))
        pid = pair2pid.get(key) or ctrl2pid.get(key[0])
        if pid is not None and cost_by_pid.get(pid) is not None:
            return cost_by_pid[pid]
        if '/' in pn:
            m = re.match(r'(\d+)', pn.rsplit('/', 1)[1].strip())
            if m:
                return int(m.group(1))
        p = pn.strip()
        if p.isdigit() and len(p) <= 6:
            return int(p)
        return hist_pair.get(key)

    return resolve


def build_sheet(kpi_file, sheet_name, head6, data, resolve):
    """テンプレート(直近月シート)をコピーして新しい月シートを構築する"""
    wb = load_workbook(kpi_file)
    if sheet_name in wb.sheetnames:
        raise SystemExit(f'シート「{sheet_name}」は既に存在します。手動で削除してから再実行してください。')
    template = wb.sheetnames[-1]
    ws = wb.copy_worksheet(wb[template])
    ws.title = sheet_name
    ws.freeze_panes = None

    # テンプレートのデータ行数を数え、今月の行数に合わせて増減する
    t_n = 0
    while ws.cell(8 + t_n, 3).value is not None:
        t_n += 1
    n = len(data)
    if n < t_n:
        ws.delete_rows(8, t_n - n)
    elif n > t_n:
        ws.insert_rows(9, n - t_n)
        for r in range(9, 9 + (n - t_n)):
            for col in range(1, 16):
                ws.cell(r, col)._style = ws.cell(8, col)._style

    for i, r in enumerate(head6, start=1):
        ws.cell(i, 1).value = r[0] if r else None
        ws.cell(i, 2).value = r[1] if len(r) > 1 else None

    last = 7 + n
    missing = []
    for i, r in enumerate(data):
        row = 8 + i
        for col in range(10):
            ws.cell(row, col + 1).value = conv(r[col]) if col < len(r) else None
        cost = resolve(r[2], r[3], r[4])
        lc = ws.cell(row, 12)
        lc.value = int(cost) if isinstance(cost, float) and cost == int(cost) else cost
        lc.fill = PatternFill(fill_type=None) if cost is not None else YELLOW
        if cost is None:
            missing.append((row, r[2], r[4]))
        ws.cell(row, 11).value = f'=SUM(I{row}*0.15)'
        ws.cell(row, 13).value = f'=SUM(H{row}*L{row})'
        ws.cell(row, 14).value = f'=SUM(I{row}-K{row}-M{row})'
        ws.cell(row, 15).value = f'=IF(I{row}=0,"",N{row}/I{row})'

    t = last + 1
    for col in 'HIJKMN':
        ws[f'{col}{t}'] = f'=SUM({col}7:{col}{last})'
    ws[f'O{t}'] = f'=IF(I{t}=0,"",N{t}/I{t})'

    e0 = t + 2
    for j, lab in enumerate(['広告費', 'ポイント費用', 'クーポン利用額', 'クーポン利用手数料']):
        row = e0 + j
        ws.cell(row, 13).value = lab
        nc = ws.cell(row, 14)
        nc.value = None
        nc.fill = YELLOW
        ws.cell(row, 15).value = f'=SUM(N{row}/I{t})' if j < 3 else None
    tot1 = e0 + 4
    ws.cell(tot1, 13).value = '合計'
    ws.cell(tot1, 14).value = f'=SUM(N{e0}:N{e0 + 3})'
    ws.cell(tot1, 15).value = f'=SUM(N{tot1}/I{t})'

    s0 = tot1 + 2
    for j, lab in enumerate(['送料', '梱包資材']):
        row = s0 + j
        ws.cell(row, 13).value = lab
        nc = ws.cell(row, 14)
        nc.value = None
        nc.fill = YELLOW
        ws.cell(row, 15).value = f'=SUM(N{row}/I{t})'
    tot2 = s0 + 2
    ws.cell(tot2, 13).value = '合計'
    ws.cell(tot2, 14).value = f'=SUM(N{s0}:N{s0 + 1})'
    ws.cell(tot2, 15).value = f'=SUM(O{s0}:O{s0 + 1})'

    g = tot2 + 2
    ws.cell(g, 13).value = '限界利益'
    ws.cell(g, 14).value = f'=SUM(N{t}-N{tot1}-N{tot2})'
    ws.cell(g + 1, 13).value = '限界利益率'
    ws.cell(g + 1, 14).value = f'=IF(I{t}=0,"",N{g}/I{t})'

    for row in ws.iter_rows(min_row=g + 2, max_row=ws.max_row):
        for c in row:
            c.value = None

    wb.save(kpi_file)
    return t, missing


def recalc_via_excel(path):
    script = f'''
set p to POSIX file "{path}"
with timeout of 600 seconds
tell application "Microsoft Excel"
    set wasRunning to running
    open p
    delay 2
    calculate
    save active workbook
    close active workbook saving no
    if not wasRunning then quit
end tell
end timeout
return "ok"
'''
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=660)
    if r.returncode != 0:
        raise RuntimeError(f'Excel再計算に失敗: {r.stderr}')


def verify(kpi_file, sheet_name, data, total_row):
    """再計算後の合計をCSVと突合し、数式エラーを検査する"""
    wb = load_workbook(kpi_file, data_only=True)
    ws = wb[sheet_name]
    exp = [sum(int(r[c]) for r in data) for c in (7, 8, 9)]
    got = [ws.cell(total_row, c).value for c in (8, 9, 10)]
    ok = got == exp
    errs = [c.coordinate for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('#')]
    return ok, exp, got, errs, ws.cell(total_row, 14).value, ws.cell(total_row, 15).value


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    csv_path = sys.argv[1]
    fname = os.path.basename(csv_path)
    if len(sys.argv) >= 3:
        sheet_name = sys.argv[2]
    else:
        m = re.match(r'\d{4}(\d{2})_', fname)
        if not m:
            raise SystemExit('シート名を自動判定できません。第2引数で指定してください(例: 8月)')
        sheet_name = f'{int(m.group(1))}月'

    kpi_file = os.environ.get('KPI_FILE_OVERRIDE', KPI_FILE)

    if kpi_file == KPI_FILE:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        from datetime import date
        base = os.path.splitext(os.path.basename(kpi_file))[0]
        bak = f'{BACKUP_DIR}/{base}_backup_{date.today():%Y%m%d}_{sheet_name}生成前.xlsx'
        import shutil
        shutil.copy2(kpi_file, bak)
        print(f'バックアップ: {os.path.basename(bak)}')

    head6, data = read_rms_csv(csv_path)
    print(f'{sheet_name}: CSVデータ {len(data)}行')

    wb_values = load_workbook(kpi_file, data_only=True)
    resolve = build_cost_resolver(wb_values)
    total_row, missing = build_sheet(kpi_file, sheet_name, head6, data, resolve)
    print(f'シート生成完了(仕入値未解決 {len(missing)}行)')
    for m_ in missing:
        print('  黄色:', m_)

    print('Excelで再計算中...')
    recalc_via_excel(kpi_file)

    ok, exp, got, errs, n_total, o_total = verify(kpi_file, sheet_name, data, total_row)
    print(f'検証: 合計{"一致" if ok else "不一致!"} (個数/売上/件数 CSV={exp} シート={got})')
    print(f'数式エラー: {len(errs)}件 {errs[:10] if errs else ""}')
    print(f'粗利: {n_total:,.0f}円 ({o_total * 100:.1f}%)' if n_total else '')
    if not ok or errs:
        sys.exit('*** FAIL — シートを確認してください ***')
    print('PASS。経費(黄色セル)入力後に限界利益が確定します。')
    print('新商品があれば: python3 sync_cost_master.py register ' + sheet_name)


if __name__ == '__main__':
    main()
